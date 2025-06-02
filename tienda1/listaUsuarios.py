import openpyxl
import os
from customtkinter import *
from openpyxl import Workbook, load_workbook

archivo = "datosUsuarios.xlsx"
def verificarExcel():
    if os.path.exists(archivo):
        wb = openpyxl.load_workbook(archivo)
        hoja = wb.active
    else:
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = "datos"
        hoja["A1"]="Usuario"
        hoja["B1"]="Contraseña"
        hoja["C1"]="Dinero"
        wb.save("datosUsuarios.xlsx") #Guardalo guardarlo
        
def agregarUsuarios(usuario, contra, ventanaReg):
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active
    usuariosExistentes= [celda.value for celda in hoja['A'][1:] if celda.value]
    #print(usuariosExistentes)
    print(usuario)
    #Lee todos los valores de los usuarios en esa columna 
    if usuario in usuariosExistentes:
        CTkLabel(ventanaReg, text="Usuario ya existente", font=("Arial", 14)).place(x=130,y=145)
    else:
        filaVacia = hoja.max_row +1 #Buscamos la fila vacia mas cercana
        hoja.cell(row=filaVacia, column=1, value=usuario)
        hoja.cell(row=filaVacia, column=2, value=contra)
        hoja.cell(row=filaVacia, column=3, value=12000)
        print("Usuario registrado")
        ventanaReg.destroy() #Cerramos la ventana
        
        filas=hoja.max_row
        columnas=hoja.max_column
        print("Filas:", filas, "Columnas:", columnas)

        wb.save("datosUsuarios.xlsx")

        #Ver en pantalla
        wb2 = load_workbook("datosUsuarios.xlsx")
        #print(wb2.sheetnames)
        hoja2 = wb2["datos"]

        print("\nLista de usuarios registrados:")
        for fila in hoja2.iter_rows(values_only=True):
            print(fila)


def diccioUsuarios(archivo="datosUsuarios.xlsx"):
    #Pasa los usuarios de excel a un diccionario
    usuarios = {}

    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active

    for fila in hoja.iter_rows(min_row=2, values_only=True): #Sin contar la primera linea
        if fila[0]:  # Solo si hay un nombre de usuario
            nombre, contra, dinero = fila
            usuarios[nombre] = {
                "contraseña": str(contra),
                "dinero": dinero
            }
    return usuarios
