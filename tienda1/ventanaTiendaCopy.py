#Tienda#
from customtkinter import *
from PIL import Image, ImageTk
import openpyxl
from pythonds3 import *

#------

def mostrarVentana(ventanaPadre):  
    valorRot=[None]          
    def quicksortPrecios(inventario):  
        quicksortAux(inventario, 0, len(inventario)-1)
        return inventario

    def quicksortAux(lista, inicio, fin):
        if inicio < fin:
            pivote = particion(lista, inicio, fin)
            quicksortAux(lista, inicio, pivote-1)
            quicksortAux(lista, pivote+1, fin)

    def particion(lista, inicio, fin):
        pivote = lista[inicio]
        izquierda = inicio + 1
        derecha = fin

        bandera = False
        while not bandera:
            while izquierda <= derecha and lista[izquierda][1] <= pivote[1]:
                izquierda = izquierda + 1
            while derecha >= izquierda and lista[derecha][1] >= pivote[1]:
                derecha = derecha - 1
            if derecha < izquierda:
                bandera = True
            else:
                lista[izquierda], lista[derecha] = lista[derecha], lista[izquierda]

        lista[inicio], lista[derecha] = lista[derecha], lista[inicio]
        return derecha

    def imprimirInventario(inventario,menu):
        fila=0
        for data in inventario:
            etiqueta = CTkLabel(master=menu, text=data[0],font=("Matura MT Script Capitals",25),text_color="#ee76ab" )
            etiqueta2 = CTkLabel(master=menu, text=data[1],font=("Matura MT Script Capitals",25),text_color="#ee76ab")
            etiqueta.grid(row=fila, column=0,padx=50)
            etiqueta2.grid(row=fila, column=1)
            fila+=1
    def eliminarEtiquetas(menu):
        for widget in menu.winfo_children():
            if isinstance(widget,CTkLabel):
                widget.destroy()
            
    def mostrarInventario():
        mostrarInventario=[]
        archivo=openpyxl.load_workbook("recetario.xlsx")
        
        ingredientes = archivo["Ingredientes"]
        
        fila=0
        for i in range(2, ingredientes.max_row):
            
            celda="A"+str(i)
            celda2="B"+str(i)
            
            valor=ingredientes[celda].value
            valor2=ingredientes[celda2].value
     
            mostrarInventario.append([valor,valor2])
        return mostrarInventario
    
#Funciones de cola -----------------------------------------------------
    
    def crearColaInventario(auxInventario):
        colaComprar=Deque()        
        for i in auxInventario:
            producto= i[0]
            colaComprar.add_rear(producto)
        return colaComprar
    
    def rotarColaIzq(colaRotar):
        elemento = colaRotar.remove_front()
        colaRotar.add_rear(elemento)
        return colaRotar, elemento

    def rotarColaDer(colaRotar):
        elemento = colaRotar.remove_rear()
        colaRotar.add_front(elemento)
        return colaRotar,elemento
    
    def imprimirCola(miCola,instruccion):
        eliminarEtiquetas(frameCola)
        if instruccion=='d':
                _,valor=rotarColaDer(miCola)
        elif instruccion=='i':
                _,valor=rotarColaIzq(miCola)
        valorCola = CTkLabel(master=frameCola,
                             text=valor,
                             font=("Matura MT Script Capitals",22),text_color="#ee76ab")
        valorCola.grid(row=0, column=0,pady=20,sticky="ew")
        valorRot[0]=valor
        return miCola

    def seleccionar(inventario,colaInventario):
        print(valorRot[0]) #ValorRot[0] es el que se imprime en frameCola
        for elemento in inventario:
            if elemento[0] == valorRot[0]:
                print(elemento)
                archivo = open("carrito.xlsx","x")
                carrito = openpyxl.load_workbook("carrito.xlsx")
                for fila in carrito.iter_rows:
                    for celda in fila:
                        if celda.value == elemento[0]:
                            carrito.cell(row=fila,column=3, value=value+1)
                            
                        else: #Si no existe aun en la hoja#
                            carrito.cell(row=fila+1,column=1, value= elemento[0])
                            carrito.cell(row=fila+1,column=2, value= elemento[1])
                            carrito.cell(row=fila+1,column=3, value= value+1)        
                
#Ventana tienda -------------------------------------------------    
    ventanaTienda = CTk()
    ventanaTienda.geometry("1300x720") ##tamaño
    ventanaTienda.title("**COMPRAR:") ##Titulo
    ventanaTienda.resizable(True,True) ##redimensionar
    ventanaTienda.configure(fg_color= "#d62957")

    ventanaTienda.grid_columnconfigure((1,2),weight=1)
    ventanaTienda.grid_rowconfigure((0),weight=1)

    ventanaTienda.iconbitmap("moneda.ico") ##icono
#--------------------------------------------------------------------
    tiendita = CTkLabel(ventanaTienda, text= "TIENDITA MAGIKA",
             text_color="#ffbfdb",
             bg_color="transparent",
             font=("Old English Text MT",50))
    tiendita.grid(row=0,column=1,padx=20,pady=5)
    
    ##Frame menu de ingredientes---------------------------------
    menu = CTkFrame(master = ventanaTienda,
                    width=550,
                    fg_color="#572e4f")
    menu.grid(row=0,column=1, padx=10,pady= 150,sticky="nsew")
    
    menu.grid_columnconfigure((1,2),weight=1)
    menu.grid_rowconfigure((1),weight=1)

#Frame menu2----------------------------------------------------
    menu2 = CTkFrame(master = ventanaTienda,
                    width=600,
                    fg_color="#8e7e9a")
    menu2.grid(row=0,column=2, padx=10, pady=30, sticky="nsew")

    menu.grid_columnconfigure((0),weight=1)
    menu2.grid_rowconfigure((1,2,3,4),weight=1)
    
    #Frame cola---------------------------------------------
    frameCola = CTkFrame(master=menu2,
                    height=25,
                    fg_color="#572e4f")
    frameCola.grid(column=0,row=1,padx=10,sticky="nsew")
  
    #Frame botones----------------------------------
    frameColaBut = CTkFrame(master=menu2,
                    height=30,
                    fg_color="#572e4f")
    frameColaBut.grid(column=0,row=2,pady=20,sticky="nsew")

    frameColaBut.grid_rowconfigure((0),weight=0)
    frameColaBut.grid_columnconfigure((1,2,3,4),weight=1)
    
    # Frame carrito-----------------------------------------
    frameCarrito= CTkFrame(master=menu2,
                           height=400,
                           fg_color="#c7cae7")
    frameCarrito.grid(column=0,row=3,pady=5,sticky="nsew")
    
    frameCarrito.grid_rowconfigure((0),weight=0)
    frameCarrito.grid_columnconfigure((1),weight=1)
    
    # Frame botonesCarrito-----------------------------------------
    frameCarritoBut= CTkFrame(master=menu2,
                              height=30,
                              fg_color="#572e4f")
    frameCarritoBut.grid(column=0,row=4,pady=5,sticky="ew")
    
    frameCarritoBut.grid_rowconfigure((0),weight=0)
    frameCarritoBut.grid_columnconfigure((1),weight=1)
    
    #Mostrar inventario------
    inventario = []
    inventario = mostrarInventario()
    imprimirInventario(inventario,menu)
    
    inventarioOrdenado=[]
    inventarioOrdenado=quicksortPrecios(inventario)
    colaInventario = crearColaInventario(inventarioOrdenado)
       
    
    #Botones ------------------------------------------
    
    ingredientesMenu=CTkButton(frameColaBut, text="Ordenar/$",
                            font=("Old English Text MT",30),
                            fg_color="#c66509",
                            command= lambda: [quicksortPrecios(inventario),eliminarEtiquetas(menu), imprimirInventario(inventario,menu),ingredientesMenu.destroy()])
    ingredientesMenu.grid(row=0,column=2, padx=1,sticky="ew")
    
    
    rotarAbajo=CTkButton(frameColaBut, text="▽",
                            font=("Old English Text MT",30),
                            fg_color="#c66509",
                            command = lambda: [valor:=imprimirCola(colaInventario,"i")])
    rotarAbajo.grid(row=0,column=1,padx=1,sticky="ew")
    
    rotarArriba=CTkButton(frameColaBut, text="△",
                            font=("Old English Text MT",30),
                            fg_color="#c66509",
                            height=25,
                            command = lambda: [valor:=imprimirCola(colaInventario,"d")])
    rotarArriba.grid(row=0,column=4,padx=1,sticky="ew")
    
    selecProducto=CTkButton(frameColaBut, text="Seleccionar",
                            font=("Old English Text MT",25),
                            fg_color="#c66509",
                            command = lambda:[seleccionar(inventario,colaInventario)])

    selecProducto.grid(row=0,column=3, padx=1,sticky="ew")
 
#-----Comportamiento al cerrar:-------------------------------------------------------------------------------
    ventanaTienda.protocol("WM_DELETE_WINDOW", lambda: [ventanaTienda.destroy(),ventanaPadre.deiconify()])
    ventanaTienda.mainloop()
    return ventanaTienda
# Ultima rotación
valor = ''

